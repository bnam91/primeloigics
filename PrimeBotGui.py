import tkinter as tk
from tkinter import ttk
import recall  # recall.py 모듈 임포트
import shipping_self  # shipping_self.py 모듈 임포트
import shutil
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QDate
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget, 
                            QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, 
                            QLineEdit, QPushButton, QTextEdit, QFileDialog,
                            QDateEdit)

class 프라임봇App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("출고 관리 시스템")
        self.geometry("1000x750")  # 화면 크기 증가
        
        # 창 최소 크기 설정
        self.minsize(1200, 860)
        
        # 전체 화면 상태 추적 변수
        self.전체화면_활성 = False
        
        # 상단 메뉴 바 생성 - 높이 40에서 30으로 줄임
        self.상단_메뉴바 = tk.Frame(self, bg="#e0e0e0", height=30)
        self.상단_메뉴바.pack(side="top", fill="x")
        
        # 메뉴 버튼 스타일 - 높이를 2에서 1로 줄임
        메뉴_버튼_스타일 = {"width": 15, "height": 1, "relief": "flat", "bg": "#e0e0e0", 
                      "activebackground": "#c0c0c0", "font": ("Helvetica", 10)}
        
        # 메뉴 버튼 생성 - pady도 5에서 3으로 줄임
        수기출고_버튼 = tk.Button(self.상단_메뉴바, text="수기출고", 
                            command=lambda: self.기능화면_표시("수기출고"), **메뉴_버튼_스타일)
        수기출고_버튼.pack(side="left", padx=5, pady=3)
        
        로켓출고_버튼 = tk.Button(self.상단_메뉴바, text="로켓출고", 
                            command=lambda: self.기능화면_표시("로켓출고"), **메뉴_버튼_스타일)
        로켓출고_버튼.pack(side="left", padx=5, pady=3)
        
        반품하기_버튼 = tk.Button(self.상단_메뉴바, text="반품하기", 
                            command=lambda: self.기능화면_표시("반품하기"), **메뉴_버튼_스타일)
        반품하기_버튼.pack(side="left", padx=5, pady=3)
        
        입고요청_버튼 = tk.Button(self.상단_메뉴바, text="입고요청", 
                            command=lambda: self.기능화면_표시("입고요청"), **메뉴_버튼_스타일)
        입고요청_버튼.pack(side="left", padx=5, pady=3)
        
        # 메뉴 버튼 참조 저장 (활성화된 탭 표시용)
        self.메뉴_버튼 = {
            "수기출고": 수기출고_버튼,
            "로켓출고": 로켓출고_버튼,
            "반품하기": 반품하기_버튼,
            "입고요청": 입고요청_버튼
        }
        
        # 메인 컨테이너 설정
        self.container = tk.Frame(self)
        self.container.pack(side="top", fill="both", expand=True)
        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=1)
        
        # 화면 전환을 위한 프레임 딕셔너리
        self.frames = {}
        
        # 상태 바 프레임 (크기 조절 버튼 포함)
        상태_프레임 = tk.Frame(self)
        상태_프레임.pack(side="bottom", fill="x")
        
        # 상태 메시지 표시 라벨
        self.상태_라벨 = tk.Label(상태_프레임, text="시스템 준비 완료", height=2)
        self.상태_라벨.pack(side="left", fill="x", expand=True, pady=5)
        
        # 초기 화면 생성 및 표시 (기본값: 수기출고)
        self.기능화면_표시("수기출고")
        
        # 키보드 단축키 바인딩 (F11 = 전체화면)
        self.bind("<F11>", lambda event: self.전체화면_토글())
        self.bind("<Escape>", lambda event: self.전체화면_해제())
    
    def 전체화면_토글(self):
        """전체화면 모드 전환"""
        if self.전체화면_활성:
            self.전체화면_해제()
        else:
            self.attributes("-fullscreen", True)
            self.전체화면_활성 = True
    
    def 전체화면_해제(self):
        """전체화면 모드 해제"""
        self.attributes("-fullscreen", False)
        self.전체화면_활성 = False
    
    def 메뉴_버튼_업데이트(self, 활성화_기능):
        """선택된 메뉴 버튼 활성화 표시"""
        # 모든 버튼 기본 스타일로 재설정
        for 기능, 버튼 in self.메뉴_버튼.items():
            if 기능 == 활성화_기능:
                # 선택된 버튼 파란색 볼드 텍스트로 강조
                버튼.config(bg="#e0e0e0", fg="blue", font=("Helvetica", 10, "bold"))
            else:
                # 기본 스타일
                버튼.config(bg="#e0e0e0", fg="black", font=("Helvetica", 10))
    
    def 기능화면_표시(self, 기능):
        # 모든 프레임 감추기
        for frame in self.frames.values():
            frame.grid_remove()
        
        # 메뉴 버튼 업데이트
        self.메뉴_버튼_업데이트(기능)
        
        # 선택된 기능에 따라 화면 표시
        if 기능 == "수기출고":
            # 수기출고 프레임이 없으면 생성
            if "수기출고" not in self.frames:
                수기출고_프레임 = tk.Frame(self.container)
                수기출고_프레임.grid(row=0, column=0, sticky="nsew")
                
                # 페이지 제목
                제목_프레임 = tk.Frame(수기출고_프레임, bg="#f0f0f0")
                제목_프레임.pack(fill="x", pady=5)
                
                제목 = tk.Label(제목_프레임, text="수기출고 테이블", 
                            font=("Helvetica", 12, "bold"), bg="#f0f0f0")
                제목.pack(padx=20, pady=5)
                
                # 수기출고 테이블 생성 함수 호출
                shipping_self.수기출고_프레임_생성(수기출고_프레임, self.상태_업데이트)
                
                self.frames["수기출고"] = 수기출고_프레임
            
            self.상태_업데이트("수기출고 테이블이 로드되었습니다.")
            self.frames["수기출고"].grid()
            self.frames["수기출고"].tkraise()
            
        elif 기능 == "로켓출고":
            if "로켓출고" not in self.frames:
                로켓출고_프레임 = tk.Frame(self.container)
                로켓출고_프레임.grid(row=0, column=0, sticky="nsew")
                
                # 페이지 제목
                제목_프레임 = tk.Frame(로켓출고_프레임, bg="#f0f0f0")
                제목_프레임.pack(fill="x", pady=5)
                
                제목 = tk.Label(제목_프레임, text="로켓출고", 
                            font=("Helvetica", 12, "bold"), bg="#f0f0f0")
                제목.pack(padx=20, pady=5)
                
                # 로켓출고 화면 내용
                내용 = tk.Label(로켓출고_프레임, text="로켓출고 기능은 아직 개발 중입니다.", font=("Helvetica", 12))
                내용.pack(pady=100)
                
                self.frames["로켓출고"] = 로켓출고_프레임
            
            self.상태_업데이트("로켓출고 화면이 로드되었습니다.")
            self.frames["로켓출고"].grid()
            self.frames["로켓출고"].tkraise()
            
        elif 기능 == "반품하기":
            if "반품하기" not in self.frames:
                반품하기_프레임 = tk.Frame(self.container)
                반품하기_프레임.grid(row=0, column=0, sticky="nsew")
                
                # 페이지 제목
                제목_프레임 = tk.Frame(반품하기_프레임, bg="#f0f0f0")
                제목_프레임.pack(fill="x", pady=5)
                
                제목 = tk.Label(제목_프레임, text="반품처리", 
                            font=("Helvetica", 12, "bold"), bg="#f0f0f0")
                제목.pack(padx=20, pady=5)
                
                # 반품하기 화면 생성
                recall.반품처리_프레임_생성(반품하기_프레임)
                
                self.frames["반품하기"] = 반품하기_프레임
            
            self.상태_업데이트("반품처리 화면이 로드되었습니다.")
            self.frames["반품하기"].grid()
            self.frames["반품하기"].tkraise()
            
        elif 기능 == "입고요청":
            if "입고요청" not in self.frames:
                # 새로 구현된 입고요청 화면 생성 함수 호출
                self.frames["입고요청"] = self.create_stock_in_request_tab()
            
            self.상태_업데이트("입고요청 화면이 로드되었습니다.")
            self.frames["입고요청"].grid()
            self.frames["입고요청"].tkraise()
    
    def 상태_업데이트(self, 메시지):
        """상태 메시지 업데이트"""
        self.상태_라벨.config(text=메시지)

    def create_stock_in_request_tab(self):
        stock_in_request_tab = tk.Frame(self.container)
        stock_in_request_tab.grid(row=0, column=0, sticky="nsew")
        
        # 페이지 제목
        제목_프레임 = tk.Frame(stock_in_request_tab, bg="#f0f0f0")
        제목_프레임.pack(fill="x", pady=5)
        
        제목 = tk.Label(제목_프레임, text="입고요청서 생성", 
                    font=("Helvetica", 12, "bold"), bg="#f0f0f0")
        제목.pack(padx=20, pady=5)
        
        # 입력 폼 프레임
        form_layout = tk.Frame(stock_in_request_tab)
        form_layout.pack(fill="x", pady=20, padx=30)
        
        # 레이블과 입력 필드 함께 배치
        tk.Label(form_layout, text="상품명:").grid(row=0, column=0, sticky="e", padx=5, pady=10)
        self.product_name_input = tk.Entry(form_layout, width=40)
        self.product_name_input.grid(row=0, column=1, sticky="w", padx=5, pady=10)
        
        tk.Label(form_layout, text="입고예정수량:").grid(row=1, column=0, sticky="e", padx=5, pady=10)
        self.expected_quantity_input = tk.Entry(form_layout, width=20)
        self.expected_quantity_input.grid(row=1, column=1, sticky="w", padx=5, pady=10)
        
        tk.Label(form_layout, text="입고예정일(YYYY-MM-DD):").grid(row=2, column=0, sticky="e", padx=5, pady=10)
        self.expected_date_input = tk.Entry(form_layout, width=20)
        self.expected_date_input.grid(row=2, column=1, sticky="w", padx=5, pady=10)
        self.expected_date_input.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # 버튼 프레임
        button_frame = tk.Frame(stock_in_request_tab)
        button_frame.pack(pady=20)
        
        # 버튼 추가
        create_button = tk.Button(button_frame, text="입고요청서 생성", command=self.create_stock_in_request,
                                 bg="#4CAF50", fg="white", font=("Helvetica", 10, "bold"), padx=15, pady=5)
        create_button.pack()
        
        # 상태 메시지 표시용 라벨
        self.stock_in_status_label = tk.Label(stock_in_request_tab, text="", fg="blue", font=("Helvetica", 10))
        self.stock_in_status_label.pack(side="bottom", fill="x", pady=10)
        
        # 안내 메시지
        info_frame = tk.Frame(stock_in_request_tab, bg="#f0f0f0", bd=1, relief="solid")
        info_frame.pack(fill="x", padx=30, pady=20)
        
        info_text = """
        입고요청서 사용 안내:
        1. 상품명과 입고예정수량을 입력합니다.
        2. 입고예정일은 YYYY-MM-DD 형식으로 입력합니다.
        3. '입고요청서 생성' 버튼을 클릭하면 당일 날짜로 파일이 생성됩니다.
        4. 생성된 파일명은 'YYMMDD_라라스토어_입고요청서.xlsx' 형식입니다.
        """
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg="#f0f0f0", padx=10, pady=10)
        info_label.pack()
        
        return stock_in_request_tab

    def create_stock_in_request(self):
        product_name = self.product_name_input.get()
        expected_quantity = self.expected_quantity_input.get()
        expected_date = self.expected_date_input.get()
        
        if not product_name or not expected_quantity:
            self.stock_in_status_label.config(text="상품명과 입고예정수량을 입력해주세요.")
            return
        
        try:
            # 현재 날짜로 파일명 생성
            current_date = datetime.now().strftime("%y%m%d")
            output_filename = f"{current_date}_라라스토어_입고요청서.xlsx"
            
            # 샘플 파일 경로 (이 파일이 프로젝트에 있어야 함)
            sample_file = "form_sample\\(샘플)250306_라라스토어_입고요청서.xlsx"
            
            # 샘플 파일 복사
            shutil.copy(sample_file, output_filename)
            
            # 엑셀 파일 열기
            workbook = openpyxl.load_workbook(output_filename)
            sheet = workbook.active
            
            # 입고예정일 입력 (D3)
            sheet['D3'] = f"입고예정일: {expected_date}"
            
            # 상품명 입력 (B6)
            sheet['B6'] = product_name
            
            # 입고예정수량 입력 (C6)
            sheet['C6'] = expected_quantity
            
            # 파일 저장
            workbook.save(output_filename)
            
            self.stock_in_status_label.config(text=f"입고요청서가 생성되었습니다: {output_filename}")
            
        except Exception as e:
            self.stock_in_status_label.config(text=f"오류 발생: {str(e)}")

# 메인 함수
if __name__ == "__main__":
    app = 프라임봇App()
    app.mainloop()