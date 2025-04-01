import chardet
import tkinter as tk
from tkinter import filedialog
import os
import pandas as pd
import openpyxl

def analyze_encoding():
    # 파일 선택 대화 상자 열기
    file_path = filedialog.askopenfilename(title="파일 선택", 
                                          filetypes=[("CSV 파일", "*.csv"), 
                                                    ("Excel 파일", "*.xlsx;*.xls"), 
                                                    ("모든 파일", "*.*")])
    
    if file_path:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # CSV 파일인 경우
        if file_ext == '.csv':
            # 파일을 바이너리 모드로 읽고 인코딩 추정
            with open(file_path, 'rb') as f:
                result = chardet.detect(f.read(10000))  # 처음 10000바이트만 검사
                print(f"인코딩 분석 결과: 인코딩: {result['encoding']}, 신뢰도: {result['confidence']}")
                print(result)
        
        # Excel 파일인 경우
        elif file_ext in ['.xlsx', '.xls']:
            try:
                # 기본 정보 분석 (pandas 사용)
                df = pd.read_excel(file_path)
                print(f"엑셀 파일 분석 결과:")
                print(f"행 수: {len(df)}")
                print(f"열 수: {len(df.columns)}")
                print(f"열 이름: {', '.join(df.columns)}")
                
                # 수식 분석 (.xlsx 파일만 지원)
                if file_ext == '.xlsx':
                    print("\n수식 분석 결과:")
                    try:
                        workbook = openpyxl.load_workbook(file_path, data_only=False)
                        
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            formula_count = 0
                            print(f"\n시트 '{sheet_name}' 수식 정보:")
                            
                            for row in sheet.iter_rows():
                                for cell in row:
                                    # 여러 방식으로 수식 접근 시도
                                    formula = None
                                    try:
                                        if hasattr(cell, 'formula') and cell.formula:
                                            formula = cell.formula
                                        elif hasattr(cell, '_formula') and cell._formula:
                                            formula = cell._formula
                                        elif hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                                            formula = cell.value
                                    except:
                                        continue
                                        
                                    if formula:
                                        formula_count += 1
                                        print(f"셀 {cell.coordinate}: {formula}")
                                        # 모든 수식 출력 시 너무 많을 수 있으므로 최대 10개만 표시
                                        if formula_count >= 10:
                                            print("...(이하 생략)...")
                                            break
                                if formula_count >= 10:
                                    break
                            
                            if formula_count == 0:
                                print("수식이 없습니다.")
                            else:
                                print(f"총 수식 수: {formula_count}개 이상")
                    except Exception as e:
                        print(f"수식 분석 중 오류 발생: {str(e)}")
                
                print(f"Excel 파일 분석 완료: {file_path}")
            except Exception as e:
                print(f"오류: 엑셀 파일 분석 중 오류 발생: {str(e)}")
        
        else:
            print("알림: 지원하지 않는 파일 형식입니다.")

# GUI 설정
root = tk.Tk()
root.title("파일 인코딩 및 엑셀 분석기")

# 버튼 추가
button = tk.Button(root, text="파일 선택 및 분석", command=analyze_encoding)
button.pack(pady=20)

# GUI 실행
root.mainloop()
