import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import shutil
import os
from datetime import datetime

# 샘플 파일 경로 설정
SAMPLE_FILE_PATH = os.path.join("form_sample", "(샘플)250306_라라스토어_입고요청서.xlsx")

def create_sample_file(filename=SAMPLE_FILE_PATH):
    """샘플 입고요청서 파일 생성"""
    # form_sample 디렉토리가 없으면 생성
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # 제목 설정
    sheet['A1'] = "입고요청서"
    sheet.merge_cells('A1:F1')
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 입고예정일 레이블
    sheet['A3'] = "입고예정일:"
    sheet['A3'].alignment = Alignment(horizontal='right', vertical='center')
    
    # 입고예정일 입력 영역
    sheet.merge_cells('D3:F3')
    sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 테이블 헤더
    headers = ["번호", "상품명", "입고예정수량", "비고"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # 첫 번째 행 번호 입력
    sheet['A6'] = 1
    sheet['A6'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 셀 너비 조정
    sheet.column_dimensions['A'].width = 7
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    
    # 파일 저장
    wb.save(filename)
    return filename

def create_stock_in_request(product_name, expected_quantity, expected_date):
    """
    입고요청서 생성
    
    매개변수:
    - product_name: 상품명
    - expected_quantity: 입고예정수량
    - expected_date: 입고예정일(YYYY-MM-DD 형식)
    
    반환값:
    - (성공 여부, 결과 메시지/파일명)
    """
    try:
        # 현재 날짜로 파일명 생성
        current_date = datetime.now().strftime("%y%m%d")
        output_filename = f"{current_date}_라라스토어_입고요청서.xlsx"
        
        # 샘플 파일이 존재하는지 확인
        sample_file = SAMPLE_FILE_PATH
        
        if not os.path.exists(sample_file):
            # 샘플 파일이 없으면 새로 생성
            create_sample_file()
        
        # 엑셀 워크북 새로 생성
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # 제목 설정
        sheet['A1'] = "입고요청서"
        sheet.merge_cells('A1:F1')
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 입고예정일 레이블과 입력
        sheet['A3'] = "입고예정일:"
        sheet['A3'].alignment = Alignment(horizontal='right', vertical='center')
        sheet.merge_cells('D3:F3')
        sheet['D3'] = expected_date
        sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 테이블 헤더
        headers = ["번호", "상품명", "입고예정수량", "비고"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # 데이터 입력
        sheet['A6'] = 1
        sheet['A6'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['B6'] = product_name
        sheet['C6'] = expected_quantity
        
        # 셀 너비 조정
        sheet.column_dimensions['A'].width = 7
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        
        # 파일 저장
        wb.save(output_filename)
        
        return (True, output_filename)
        
    except Exception as e:
        return (False, str(e))

# 모듈 직접 실행 시 샘플 파일 생성
if __name__ == "__main__":
    filename = create_sample_file()
    print(f"샘플 파일이 생성되었습니다: {filename}") 